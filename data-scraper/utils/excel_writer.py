"""
Excel Writer Module
===================
Firma verilerini Excel formatında kaydeden modül.

Özellikler:
- Türkçe başlıklar
- Kolon sıralaması
- Append fonksiyonu (mevcut dosyaya ekleme)
- Otomatik genişlik ayarlama
"""

import os
from datetime import datetime
from typing import Dict, List, Any, Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Config'den ayarları al
try:
    from config.settings import (
        EXCEL_FILE,
        EXCEL_COLUMNS,
        EXCEL_HEADERS,
        OUTPUT_DIR
    )
except ImportError:
    OUTPUT_DIR = 'output'
    EXCEL_FILE = os.path.join(OUTPUT_DIR, 'company_data.xlsx')
    EXCEL_COLUMNS = [
        'company_name', 'website', 'sector', 'phone', 'email',
        'address', 'catalog_count', 'catalog_files', 'status', 'scrape_date'
    ]
    EXCEL_HEADERS = {
        'company_name': 'Firma Adı',
        'website': 'Web Sitesi',
        'sector': 'Sektör',
        'phone': 'Telefon',
        'email': 'E-posta',
        'address': 'Adres',
        'catalog_count': 'Katalog Sayısı',
        'catalog_files': 'Katalog Dosyaları',
        'status': 'Durum',
        'scrape_date': 'Tarih'
    }

from utils.logger import get_logger
from utils.validators import validate_company_data

logger = get_logger(__name__)


class ExcelWriter:
    """
    Firma verilerini Excel dosyasına yazan sınıf.

    Özellikler:
    - Yeni dosya oluşturma veya mevcut dosyaya ekleme
    - Türkçe başlıklar
    - Otomatik formatlama
    - Katalog bulunamayan firmaları filtreleme

    Attributes:
        file_path: Excel dosya yolu
        columns: Kolon listesi
        headers: Kolon başlıkları (Türkçe)

    Example:
        >>> writer = ExcelWriter()
        >>> writer.append_company(company_data)
        >>> writer.save()
    """

    def __init__(self, file_path: Optional[str] = None):
        """
        ExcelWriter'ı başlatır.

        Args:
            file_path: Excel dosya yolu (varsayılan: EXCEL_FILE)
        """
        self.file_path = file_path or EXCEL_FILE
        self.columns = EXCEL_COLUMNS
        self.headers = EXCEL_HEADERS
        self.pending_data: List[Dict[str, Any]] = []

        # Output dizinini oluştur
        os.makedirs(os.path.dirname(self.file_path) or OUTPUT_DIR, exist_ok=True)

    def append_company(self, company_data: Dict[str, Any]) -> bool:
        """
        Firma verisini bekleyen listeye ekler.

        Katalog bulunamayan firmalar eklenmez!

        Args:
            company_data: Firma verisi dictionary'si

        Returns:
            bool: Veri eklendiyse True, atlandıyse False

        Example:
            >>> data = {
            ...     'company_name': 'ABC Firma',
            ...     'website': 'https://abc.com',
            ...     'catalog_count': 2,
            ...     'status': 'SUCCESS'
            ... }
            >>> writer.append_company(data)
            True
        """
        # Katalog kontrolü - bulunamayan firmalar eklenmez
        if not validate_company_data(company_data):
            company_name = company_data.get('company_name', 'Bilinmeyen')
            logger.info(f"[SKIP] {company_name} - Katalog bulunamadı, Excel'e yazılmıyor")
            return False

        # Veriyi hazırla
        prepared_data = self._prepare_data(company_data)
        self.pending_data.append(prepared_data)

        company_name = company_data.get('company_name', 'Bilinmeyen')
        logger.debug(f"[ADD] {company_name} - Listeye eklendi")

        return True

    def append_companies(self, companies: List[Dict[str, Any]]) -> int:
        """
        Birden fazla firmayı bekleyen listeye ekler.

        Args:
            companies: Firma listesi

        Returns:
            int: Eklenen firma sayısı

        Example:
            >>> count = writer.append_companies(company_list)
            >>> print(f"{count} firma eklendi")
        """
        added_count = 0

        for company in companies:
            if self.append_company(company):
                added_count += 1

        logger.info(f"Toplam {added_count}/{len(companies)} firma listeye eklendi")
        return added_count

    def save(self) -> bool:
        """
        Bekleyen verileri Excel dosyasına kaydeder.

        Mevcut dosya varsa sonuna ekler, yoksa yeni dosya oluşturur.

        Returns:
            bool: Başarılı ise True

        Example:
            >>> writer.save()
            True
        """
        if not self.pending_data:
            logger.warning("Kaydedilecek veri yok")
            return False

        try:
            # DataFrame oluştur
            df = pd.DataFrame(self.pending_data)

            # Kolon sıralaması
            df = self._reorder_columns(df)

            # Mevcut dosya varsa ekle, yoksa yeni oluştur
            if os.path.exists(self.file_path):
                self._append_to_existing(df)
            else:
                self._create_new_file(df)

            logger.info(f"Excel kaydedildi: {self.file_path} ({len(self.pending_data)} kayıt)")

            # Bekleyen veriyi temizle
            self.pending_data.clear()

            return True

        except Exception as e:
            logger.error(f"Excel kaydetme hatası: {str(e)}")
            return False

    def _prepare_data(self, company_data: Dict[str, Any]) -> Dict[str, Any]:
        """
        Firma verisini Excel formatına hazırlar.

        Args:
            company_data: Ham firma verisi

        Returns:
            Dict: Hazırlanmış veri
        """
        prepared = {}

        for col in self.columns:
            value = company_data.get(col, '')

            # catalog_files listesini string'e çevir
            if col == 'catalog_files' and isinstance(value, list):
                value = ', '.join(value) if value else ''

            # scrape_date yoksa ekle
            if col == 'scrape_date' and not value:
                value = datetime.now().strftime('%Y-%m-%d %H:%M')

            # None değerleri boş string yap
            if value is None:
                value = ''

            prepared[col] = value

        return prepared

    def _reorder_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        DataFrame kolonlarını belirtilen sıraya göre düzenler.

        Args:
            df: DataFrame

        Returns:
            pd.DataFrame: Sıralanmış DataFrame
        """
        # Mevcut kolonlar
        existing_cols = [col for col in self.columns if col in df.columns]

        # Eksik kolonları ekle
        for col in self.columns:
            if col not in df.columns:
                df[col] = ''

        return df[self.columns]

    def _create_new_file(self, df: pd.DataFrame) -> None:
        """
        Yeni Excel dosyası oluşturur.

        Args:
            df: Kaydedilecek DataFrame
        """
        # Türkçe başlıkları uygula
        df_renamed = df.rename(columns=self.headers)

        # Excel'e kaydet
        df_renamed.to_excel(self.file_path, index=False, engine='openpyxl')

        # Formatlama uygula
        self._apply_formatting()

    def _append_to_existing(self, df: pd.DataFrame) -> None:
        """
        Mevcut Excel dosyasına veri ekler.

        Args:
            df: Eklenecek DataFrame
        """
        try:
            # Mevcut dosyayı oku
            existing_df = pd.read_excel(self.file_path, engine='openpyxl')

            # Kolon isimlerini orijinale çevir (Türkçe -> İngilizce)
            reverse_headers = {v: k for k, v in self.headers.items()}
            existing_df = existing_df.rename(columns=reverse_headers)

            # Birleştir
            combined_df = pd.concat([existing_df, df], ignore_index=True)

            # Tekrar Türkçe başlıklarla kaydet
            combined_df_renamed = combined_df.rename(columns=self.headers)
            combined_df_renamed.to_excel(self.file_path, index=False, engine='openpyxl')

            # Formatlama uygula
            self._apply_formatting()

        except Exception as e:
            logger.error(f"Mevcut dosyaya ekleme hatası: {str(e)}")
            # Hata durumunda yeni dosya oluştur
            self._create_new_file(df)

    def _apply_formatting(self) -> None:
        """Excel dosyasına formatlama uygular."""
        try:
            wb = load_workbook(self.file_path)
            ws = wb.active

            # Başlık stili
            header_font = Font(bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center')

            # Border stili
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Başlık satırını formatla
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

            # Veri hücrelerine border ekle
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical='center')

            # Kolon genişliklerini ayarla
            column_widths = {
                'A': 25,  # Firma Adı
                'B': 35,  # Web Sitesi
                'C': 15,  # Sektör
                'D': 18,  # Telefon
                'E': 30,  # E-posta
                'F': 40,  # Adres
                'G': 12,  # Katalog Sayısı
                'H': 50,  # Katalog Dosyaları
                'I': 12,  # Durum
                'J': 16   # Tarih
            }

            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width

            # Satır yüksekliği
            for row in range(1, ws.max_row + 1):
                ws.row_dimensions[row].height = 20

            # Kaydet
            wb.save(self.file_path)

        except Exception as e:
            logger.warning(f"Formatlama uygulanamadı: {str(e)}")

    def get_existing_companies(self) -> List[str]:
        """
        Mevcut Excel dosyasındaki firma isimlerini döndürür.

        Duplicate kontrolü için kullanılabilir.

        Returns:
            List[str]: Firma isimleri listesi
        """
        if not os.path.exists(self.file_path):
            return []

        try:
            df = pd.read_excel(self.file_path, engine='openpyxl')

            # Türkçe başlık varsa
            if 'Firma Adı' in df.columns:
                return df['Firma Adı'].tolist()
            elif 'company_name' in df.columns:
                return df['company_name'].tolist()

            return []

        except Exception as e:
            logger.error(f"Mevcut firmalar okunamadı: {str(e)}")
            return []

    def get_stats(self) -> Dict[str, int]:
        """
        Excel dosyası istatistiklerini döndürür.

        Returns:
            Dict: İstatistikler (total_records, success, partial, etc.)
        """
        if not os.path.exists(self.file_path):
            return {'total_records': 0, 'success': 0, 'partial': 0}

        try:
            df = pd.read_excel(self.file_path, engine='openpyxl')

            status_col = 'Durum' if 'Durum' in df.columns else 'status'

            stats = {
                'total_records': len(df),
                'success': len(df[df[status_col] == 'SUCCESS']) if status_col in df.columns else 0,
                'partial': len(df[df[status_col] == 'PARTIAL']) if status_col in df.columns else 0
            }

            return stats

        except Exception as e:
            logger.error(f"İstatistikler okunamadı: {str(e)}")
            return {'total_records': 0, 'success': 0, 'partial': 0}
